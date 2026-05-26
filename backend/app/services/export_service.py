import io
import csv
import json
from typing import List
from app.models.lead import Lead
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


LEAD_FIELDS = [
    ("nome", "Nome"),
    ("categoria", "Nicho/Categoria"),
    ("status", "Status"),
    ("telefone", "Telefone"),
    ("email", "Email"),
    ("whatsapp", "WhatsApp"),
    ("instagram", "Instagram"),
    ("facebook", "Facebook"),
    ("website", "Website"),
    ("endereco", "Endereço"),
    ("cidade", "Cidade"),
    ("estado", "Estado"),
    ("nota", "Nota Google"),
    ("total_reviews", "Total Reviews"),
    ("google_maps_url", "Google Maps URL"),
    ("observacoes", "Observações"),
    ("scraped_at", "Data Captura"),
]


def leads_to_csv(leads: List[Lead]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_ALL)

    # Header
    writer.writerow([label for _, label in LEAD_FIELDS])

    # Rows
    for lead in leads:
        row = []
        for field, _ in LEAD_FIELDS:
            value = getattr(lead, field, None)
            if value is None:
                row.append("")
            elif field == "status":
                row.append(value.value if hasattr(value, "value") else str(value))
            elif field == "scraped_at":
                row.append(value.strftime("%d/%m/%Y %H:%M") if value else "")
            else:
                row.append(str(value))
        writer.writerow(row)

    return output.getvalue().encode("utf-8-sig")  # UTF-8 BOM for Excel compatibility


def leads_to_excel(leads: List[Lead]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Styles
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    # Header row
    for col, (_, label) in enumerate(LEAD_FIELDS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 30

    # Data rows
    for row_idx, lead in enumerate(leads, start=2):
        is_alt = row_idx % 2 == 0
        for col, (field, _) in enumerate(LEAD_FIELDS, start=1):
            value = getattr(lead, field, None)
            if value is None:
                display = ""
            elif field == "status":
                display = value.value if hasattr(value, "value") else str(value)
            elif field == "scraped_at":
                display = value.strftime("%d/%m/%Y %H:%M") if value else ""
            else:
                display = str(value)

            cell = ws.cell(row=row_idx, column=col, value=display)
            cell.border = thin_border
            if is_alt:
                cell.fill = alt_fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    # Auto-fit columns
    col_widths = {
        "nome": 35, "categoria": 20, "status": 15, "telefone": 18,
        "email": 35, "whatsapp": 18, "instagram": 30, "facebook": 30,
        "website": 35, "endereco": 45, "cidade": 20, "estado": 10,
        "nota": 10, "total_reviews": 12, "google_maps_url": 40,
        "observacoes": 40, "scraped_at": 18,
    }
    for col, (field, _) in enumerate(LEAD_FIELDS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(field, 15)

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto filter
    ws.auto_filter.ref = ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def leads_to_json(leads: List[Lead]) -> bytes:
    result = []
    for lead in leads:
        item = {}
        for field, label in LEAD_FIELDS:
            value = getattr(lead, field, None)
            if field == "status":
                item[field] = value.value if hasattr(value, "value") else value
            elif field == "scraped_at":
                item[field] = value.isoformat() if value else None
            else:
                item[field] = value
        result.append(item)
    return json.dumps(result, ensure_ascii=False, indent=2).encode("utf-8")
